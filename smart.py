#!/usr/bin/env python3
"""
smart.py
1. Reads UPCs from barcodes.csv (expects a 'barcode' column).
2. Queries Go‑UPC API with retry/backoff on HTTP 429.
3. Downloads each product’s image into `images/` (converting WebP to PNG when needed).
4. Writes an Excel file `product_data.xlsx` with columns:
   UPC, Product Name, Brand Name, Image URL
5. Records any UPCs that failed Go‑UPC lookup; writes those UPCs to `failed_upcs.xlsx`.
6. Queries Open Food Facts for those failed UPCs, retrieves:
   UPC, Product Name, Brands, Image URL via OFF API.
7. Downloads OFF images as well, saving into `images/`.
8. For any UPC where Go‑UPC returned partial data (missing name, brand, or image), attempts fallback to OFF API and logs the replacement.
9. Writes a separate Excel file `off_products.xlsx` for the OFF results.
10. Prints a summary of successes and failures for both APIs.

Dependencies:
    pip install requests openpyxl pillow
"""

import csv
import io
import re
import sys
import time
from pathlib import Path
from typing import List, Dict, Any, Tuple

import requests
from requests.exceptions import HTTPError, Timeout, RequestException
try:
    from PIL import Image as PILImage
except ImportError:
    print("ERROR: Pillow is required to process images. Install with 'pip install pillow'", file=sys.stderr)
    sys.exit(1)
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# --- Configuration ---
API_KEY = ''
GOUPC_BASE = 'https://go-upc.com/api/v1/code'
OFF_BASE = 'https://world.openfoodfacts.net/api/v2/product'
INPUT_CSV = Path('extractedBarcodesFromImagecsv.csv')
OUTPUT_XLSX = Path('final_product_data.xlsx')
FAILED_XLSX = Path('final_failed_upcs.xlsx')
OFF_XLSX = Path('final_off_products.xlsx')
IMAGE_DIR = Path('images')
TIMEOUT = 10
MAX_RETRIES = 5
INITIAL_WAIT = 1
REQUEST_DELAY = 0.2

# Style
RED_FILL = PatternFill(start_color='FFC7CE',
                       end_color='FFC7CE', fill_type='solid')

# --- Image filename customization ---


def make_image_filename(brand: Any, product_name: Any, ext: str) -> str:
    brand_str = str(brand) if brand else ''
    name_str = str(product_name) if product_name else ''
    safe_brand = re.sub(r'[^A-Za-z0-9 _\-]+', '_', brand_str)
    safe_brand = re.sub(r'_+', '_', safe_brand).strip('_ ')
    safe_name = re.sub(r'[^A-Za-z0-9 _\-]+', '_', name_str)
    safe_name = re.sub(r'_+', '_', safe_name).strip('_ ')
    filename = f"{safe_brand}_{safe_name}{ext}"
    filename = re.sub(r'_+', '_', filename).strip('_ ')
    return filename or f"unknown{ext}"


def lookup_upc(upc: str) -> Dict[str, Any]:
    print(f"Please Wait... querying Go‑UPC for {upc}")
    url = f"{GOUPC_BASE}/{upc}"
    wait = INITIAL_WAIT
    for attempt in range(1, MAX_RETRIES+1):
        try:
            resp = requests.get(
                url, params={'key': API_KEY, 'format': 'true'}, timeout=TIMEOUT)
            if resp.status_code == 429:
                print(
                    f"429 rate limit on Go‑UPC for {upc}, retry {attempt}/{MAX_RETRIES}")
                time.sleep(wait)
                wait *= 2
                continue
            resp.raise_for_status()
            return resp.json()
        except Timeout:
            print(f"Timeout querying Go‑UPC for {upc}")
        except HTTPError as he:
            code = he.response.status_code
            if code == 404:
                print(f"Go‑UPC: {upc} not found (404)")
                raise
            elif code in (401, 403):
                sys.exit(f"Authentication error ({code}) on Go‑UPC API")
            else:
                print(f"HTTP error {code} for {upc} on Go‑UPC")
                raise
        except RequestException as re_err:
            print(f"Network error querying Go‑UPC for {upc}: {re_err}")
        break
    raise HTTPError(f"Failed Go‑UPC lookup for {upc} after retries")


def lookup_off(upc: str) -> Dict[str, Any]:
    print(f"Please Wait... querying OFF API for {upc}")
    url = f"{OFF_BASE}/{upc}"
    params = {'fields': 'product_name,brands,image_url'}
    try:
        resp = requests.get(url, params=params, timeout=TIMEOUT)
        resp.raise_for_status()
        data = resp.json()
        if data.get('status') != 1:
            print(f"OFF API: {upc} not found (status={data.get('status')})")
            return {'code': upc, 'status': 0, 'product': {}}
        return data
    except Exception as e:
        print(f"Error querying OFF API for {upc}: {e}")
        return {'code': upc, 'status': 0, 'product': {}}


def process_all(csv_path: Path) -> Tuple[List[Dict[str, Any]], List[str]]:
    records, failed = [], []
    if not csv_path.exists():
        sys.exit(f"ERROR: Input CSV not found: {csv_path}")
    try:
        with csv_path.open(newline='', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            if 'barcode' not in (reader.fieldnames or []):
                sys.exit("ERROR: 'barcode' column missing in CSV.")
            for row in reader:
                upc = row.get('barcode', '').strip()
                if not upc:
                    print("Skipping empty barcode row")
                    continue
                try:
                    records.append(lookup_upc(upc))
                except HTTPError:
                    failed.append(upc)
                time.sleep(REQUEST_DELAY)
    except csv.Error as ce:
        sys.exit(f"CSV parse error: {ce}")
    return records, failed


def save_to_xlsx(data: List[Dict[str, Any]], output_file: Path, image_dir: Path) -> None:
    image_dir.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Products'
    ws.append(['UPC', 'Product Name', 'Brand Name', 'Image URL'])
    for col in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col].width = 30

    for idx, item in enumerate(data, start=2):
        upc = item.get('code', '')
        prod = item.get('product', {})
        name = prod.get('name', '')
        brand = prod.get('brand', '')
        url = prod.get('imageUrl', '')

        # Fallback if any field missing
        if any(not v for v in (name, brand, url)):
            off = lookup_off(upc)
            off_prod = off.get('product', {})
            if not name:
                print(f"{upc}: name missing, fetching from OFF API")
                name = off_prod.get('product_name', '') or name
            if not brand:
                print(f"{upc}: brand missing, fetching from OFF API")
                brand = off_prod.get('brands', '') or brand
            if not url:
                print(f"{upc}: image URL missing, fetching from OFF API")
                url = off_prod.get('image_url', '') or url

        # Download image
        if url:
            ext = Path(url).suffix.split('?')[0].lower() or '.jpg'
            fn = make_image_filename(brand, name, ext)
            dest = image_dir / fn
            try:
                r = requests.get(url, timeout=TIMEOUT)
                r.raise_for_status()
                raw = r.content
                if ext == '.webp':
                    img = PILImage.open(io.BytesIO(raw)).convert('RGB')
                    dest = dest.with_suffix('.png')
                    img.save(dest, 'PNG')
                else:
                    dest.write_bytes(raw)
            except Exception as e:
                print(f"Image download error for {upc}: {e}")

        row = [upc, name or 'Not Found',
               brand or 'Not Found', url or 'Not Found']
        ws.append(row)
        for c in range(1, 5):
            cell = ws.cell(row=idx, column=c)
            if not cell.value or cell.value == 'Not Found':
                cell.value = 'Not Found'
                cell.fill = RED_FILL
        ws.row_dimensions[idx].height = 90

    try:
        wb.save(output_file)
    except Exception as e:
        sys.exit(f"Error saving Excel: {e}")


def save_failed_xlsx(failed: List[str], output_file: Path) -> None:
    if not failed:
        return
    wb = Workbook()
    ws = wb.active
    ws.title = 'Failed UPCs'
    ws.append(['UPC'])
    ws.column_dimensions['A'].width = 50
    for idx, upc in enumerate(failed, start=2):
        cell = ws.cell(row=idx, column=1, value=upc)
        cell.fill = RED_FILL
        ws.row_dimensions[idx].height = 50
    wb.save(output_file)


def process_off(failed: List[str]) -> List[Dict[str, Any]]:
    return [lookup_off(u) for u in failed]


def save_off_xlsx(data: List[Dict[str, Any]], output_file: Path, image_dir: Path) -> None:
    image_dir.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = 'OFF Products'
    ws.append(['UPC', 'Product Name', 'Brands', 'Image URL'])
    for col in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col].width = 30
    for idx, item in enumerate(data, start=2):
        upc = item.get('code', '')
        prod = item.get('product', {})
        name = prod.get('product_name', '')
        brands = prod.get('brands', '')
        url = prod.get('image_url', '')
        if url:
            ext = Path(url).suffix.split('?')[0].lower() or '.jpg'
            fn = make_image_filename(brands, name, ext)
            dest = image_dir / fn
            try:
                r = requests.get(url, timeout=TIMEOUT)
                r.raise_for_status()
                raw = r.content
                if ext == '.webp':
                    img = PILImage.open(io.BytesIO(raw)).convert('RGB')
                    dest = dest.with_suffix('.png')
                    img.save(dest, 'PNG')
                else:
                    dest.write_bytes(raw)
            except Exception as e:
                print(f"OFF image download error for {upc}: {e}")
        row = [upc, name or 'Not Found',
               brands or 'Not Found', url or 'Not Found']
        ws.append(row)
        for c in range(1, 5):
            cell = ws.cell(row=idx, column=c)
            if not cell.value or cell.value == 'Not Found':
                cell.value = 'Not Found'
                cell.fill = RED_FILL
        ws.row_dimensions[idx].height = 90
    wb.save(output_file)


def main():
    try:
        products, failed = process_all(INPUT_CSV)
        print(f"Go‑UPC: {len(products)} succeeded, {len(failed)} failed.")
        save_to_xlsx(products, OUTPUT_XLSX, IMAGE_DIR)
        save_failed_xlsx(failed, FAILED_XLSX)
        off_results = process_off(failed)
        found = sum(1 for r in off_results if r.get('status') == 1)
        not_found = len(off_results) - found
        print(f"OFF API: {found} found, {not_found} not found.")
        save_off_xlsx(off_results, OFF_XLSX, IMAGE_DIR)
    except Exception as e:
        sys.exit(f"Unexpected error: {e}")


if __name__ == '__main__':
    main()
